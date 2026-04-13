"""
export.py — Excel export logic for the 13F ownership research app.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='002147', end_color='002147', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
ALT_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
DATA_FONT = Font(name='Arial', size=10)


def _write_sheet(ws, data):
    if not data:
        ws.append(['No data'])
        return

    flat_data = []
    for row in data:
        flat = {}
        for k, v in row.items():
            if k.startswith('_'):
                continue
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat[f"{k}_{k2}"] = v2
            else:
                flat[k] = v
        flat_data.append(flat)

    headers = ['#'] + list(flat_data[0].keys())
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for row_idx, record in enumerate(flat_data, 2):
        num_cell = ws.cell(row=row_idx, column=1)
        num_cell.value = row_idx - 1
        num_cell.font = DATA_FONT
        num_cell.alignment = Alignment(horizontal='right')
        for col_idx, key in enumerate(headers[1:], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = record.get(key)
            if isinstance(val, (list, dict)):
                val = str(val)
            cell.value = val
            cell.font = DATA_FONT

            if isinstance(val, (int, float)) and val is not None:
                if 'pct' in key.lower() or 'percent' in key.lower():
                    cell.number_format = '0.00"%"'
                elif 'value' in key.lower() or 'aum' in key.lower() or 'cap' in key.lower():
                    cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL

    for col_idx, key in enumerate(headers, 1):
        max_len = len(str(key))
        for row_idx in range(2, min(len(flat_data) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)


def _sanitize_sheet_name(name):
    name = str(name).replace(':', '-').replace('/', '-').replace('\\', '-')
    name = name.replace('?', '').replace('*', '').replace('[', '(').replace(']', ')')
    return name[:31] or 'Sheet'


def build_excel(data, sheet_name='Data'):
    """Build a formatted .xlsx workbook from a list of dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(sheet_name)
    _write_sheet(ws, data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_multisheet(sheets):
    """Build a multi-sheet .xlsx workbook from a dict of {sheet_name: list_of_dicts}.

    Empty sheets render a 'No data' row. Non-list values are wrapped into a single-row sheet.
    Sheet names are sanitized (Excel limit: 31 chars, no special chars).
    """
    wb = Workbook()
    first = True
    used_names = set()
    for raw_name, value in sheets.items():
        name = _sanitize_sheet_name(raw_name)
        # Uniqueness guard — truncate can collide
        original, i = name, 2
        while name in used_names:
            suffix = f'_{i}'
            name = _sanitize_sheet_name(original[:31 - len(suffix)] + suffix)
            i += 1
        used_names.add(name)

        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)

        if isinstance(value, dict):
            rows = [value]
        elif isinstance(value, list):
            rows = value
        else:
            rows = [{'value': value}]
        _write_sheet(ws, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
