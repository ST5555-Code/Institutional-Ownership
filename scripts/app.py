"""
Flask web application for 13F institutional ownership research.
Replaces Jupyter notebook for day-to-day browser-based research.
"""

import argparse
import io
import os
import shutil
import sys
from datetime import datetime

import duckdb
import pandas as pd
from flask import Flask, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Flask app setup
# ---------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, 'data', '13f.duckdb')
DB_SNAPSHOT_PATH = os.path.join(BASE_DIR, 'data', '13f_readonly.duckdb')

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'web', 'templates'),
    static_folder=os.path.join(BASE_DIR, 'web', 'static'),
)


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def _resolve_db_path():
    """Return the best available database path.

    Try the main database first (read_only=True). If it is locked by a writer
    (e.g. fetch_nport.py), fall back to a snapshot copy so the web app can
    still serve data while the pipeline runs.
    """
    try:
        con = duckdb.connect(DB_PATH, read_only=True)
        con.close()
        return DB_PATH
    except Exception:
        # Main DB is locked — use or create a snapshot
        if os.path.exists(DB_SNAPSHOT_PATH):
            return DB_SNAPSHOT_PATH
        # Create snapshot from the main DB
        try:
            print(f'  Database locked — creating read-only snapshot...')
            shutil.copy2(DB_PATH, DB_SNAPSHOT_PATH)
            print(f'  Snapshot ready: {DB_SNAPSHOT_PATH}')
            return DB_SNAPSHOT_PATH
        except Exception as e2:
            raise RuntimeError(
                f'Cannot open database (locked) and cannot create snapshot: {e2}'
            )


# Resolved once at import/startup; updated if needed
_active_db_path = None


def _init_db_path():
    """Resolve the database path at startup."""
    global _active_db_path
    _active_db_path = _resolve_db_path()


def get_db():
    """Open a read-only DuckDB connection. Caller must close it."""
    global _active_db_path
    if _active_db_path is None:
        _active_db_path = _resolve_db_path()
    try:
        return duckdb.connect(_active_db_path, read_only=True)
    except Exception:
        # The previously resolved path may have become stale; re-resolve
        _active_db_path = _resolve_db_path()
        return duckdb.connect(_active_db_path, read_only=True)


def get_cusip(con, ticker):
    """Resolve ticker to CUSIP."""
    row = con.execute(
        "SELECT cusip FROM securities WHERE ticker = ? LIMIT 1", [ticker]
    ).fetchone()
    return row[0] if row else ''


def _clean_val(v):
    """Replace NaN/Inf with None; convert numpy types to native Python."""
    if v is None:
        return None
    if isinstance(v, float):
        import math
        if math.isnan(v) or math.isinf(v):
            return None
    # numpy scalar types — convert to native Python
    try:
        import numpy as np
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating,)):
            import math
            if math.isnan(v) or math.isinf(v):
                return None
            return float(v)
        if isinstance(v, np.bool_):
            return bool(v)
    except ImportError:
        pass
    return v


def clean_for_json(data):
    """Recursively clean NaN/Inf/numpy types from dicts and lists."""
    if isinstance(data, list):
        return [{k: _clean_val(v) for k, v in row.items()} if isinstance(row, dict) else row
                for row in data]
    if isinstance(data, dict):
        return {k: clean_for_json(v) if isinstance(v, (dict, list))
                else _clean_val(v)
                for k, v in data.items()}
    return _clean_val(data)


def df_to_records(df):
    """Convert DataFrame to list of dicts with NaN/Inf replaced by None."""
    records = df.to_dict(orient='records')
    return clean_for_json(records)


# ---------------------------------------------------------------------------
# Query functions — ported exactly from research.ipynb
# ---------------------------------------------------------------------------

def query1(ticker):
    """Current shareholder register — two-level parent/fund hierarchy."""
    con = get_db()
    try:
        cusip = get_cusip(con, ticker)
        # Parent-level with distinct fund count
        parents = con.execute("""
            WITH by_fund AS (
                SELECT
                    COALESCE(h.inst_parent_name, h.manager_name) as parent_name,
                    h.fund_name,
                    h.cik,
                    COALESCE(h.manager_type, 'unknown') as type,
                    h.market_value_live,
                    h.shares,
                    h.pct_of_float
                FROM holdings h
                WHERE h.quarter = '2025Q4'
                  AND (h.ticker = ? OR h.cusip = ?)
            )
            SELECT
                parent_name,
                MAX(type) as type,
                SUM(market_value_live) as total_value_live,
                SUM(shares) as total_shares,
                SUM(pct_of_float) as pct_float,
                COUNT(DISTINCT fund_name) as child_count
            FROM by_fund
            GROUP BY parent_name
            ORDER BY total_value_live DESC NULLS LAST
            LIMIT 15
        """, [ticker, cusip]).fetchdf()

        results = []
        for rank, (_, parent) in enumerate(parents.iterrows(), 1):
            pname = parent['parent_name']
            child_count = int(parent['child_count'])
            results.append({
                'rank': rank,
                'institution': pname,
                'value_live': parent['total_value_live'],
                'shares': parent['total_shares'],
                'pct_float': parent['pct_float'],
                'type': parent['type'],
                'is_parent': child_count >= 2,
                'child_count': child_count,
                'level': 0,
            })
            # Only show child rows if 2+ distinct funds
            if child_count < 2:
                continue
            safe_name = str(pname).replace("'", "''") if pname else ''
            subs = con.execute(f"""
                SELECT
                    h.fund_name,
                    COALESCE(h.manager_type, 'unknown') as type,
                    SUM(h.market_value_live) as value_live,
                    SUM(h.shares) as shares,
                    SUM(h.pct_of_float) as pct_of_float
                FROM holdings h
                WHERE h.quarter = '2025Q4'
                  AND (h.ticker = ? OR h.cusip = ?)
                  AND COALESCE(h.inst_parent_name, h.manager_name) = '{safe_name}'
                GROUP BY h.fund_name, type
                ORDER BY value_live DESC NULLS LAST
                LIMIT 5
            """, [ticker, cusip]).fetchdf()
            for _, sub in subs.iterrows():
                results.append({
                    'rank': None,
                    'institution': sub['fund_name'],
                    'value_live': sub['value_live'],
                    'shares': sub['shares'],
                    'pct_float': sub['pct_of_float'],
                    'type': sub['type'],
                    'is_parent': False,
                    'child_count': 0,
                    'level': 1,
                })
        return results
    finally:
        con.close()


def query2(ticker):
    """4-quarter ownership change (Q1 vs Q4 2025)."""
    con = get_db()
    try:
        cusip = get_cusip(con, ticker)
        # Top 15 parents by Q4 value
        top_parents = con.execute("""
            SELECT COALESCE(inst_parent_name, manager_name) as parent_name,
                   SUM(market_value_live) as parent_val
            FROM holdings
            WHERE quarter = '2025Q4' AND (ticker = ? OR cusip = ?)
            GROUP BY parent_name
            ORDER BY parent_val DESC NULLS LAST
            LIMIT 15
        """, [ticker, cusip]).fetchdf()['parent_name'].tolist()

        q2 = con.execute("""
            WITH q1_agg AS (
                SELECT cik, manager_name,
                       COALESCE(inst_parent_name, manager_name) as parent_name,
                       MAX(manager_type) as manager_type,
                       SUM(shares) as q1_shares
                FROM holdings
                WHERE quarter = '2025Q1' AND (ticker = ? OR cusip = ?)
                GROUP BY cik, manager_name, parent_name
            ),
            q4_agg AS (
                SELECT cik, manager_name,
                       COALESCE(inst_parent_name, manager_name) as parent_name,
                       MAX(manager_type) as manager_type,
                       SUM(shares) as q4_shares
                FROM holdings
                WHERE quarter = '2025Q4' AND (ticker = ? OR cusip = ?)
                GROUP BY cik, manager_name, parent_name
            ),
            combined AS (
                SELECT
                    COALESCE(q4.parent_name, q1.parent_name) as parent_name,
                    COALESCE(q4.manager_name, q1.manager_name) as fund_name,
                    COALESCE(q4.cik, q1.cik) as cik,
                    COALESCE(q4.manager_type, q1.manager_type, 'unknown') as type,
                    q1.q1_shares,
                    q4.q4_shares,
                    COALESCE(q4.q4_shares, 0) - COALESCE(q1.q1_shares, 0) as change_shares,
                    CASE
                        WHEN q1.q1_shares > 0 AND q4.q4_shares IS NOT NULL
                        THEN ROUND((q4.q4_shares - q1.q1_shares) * 100.0 / q1.q1_shares, 1)
                        WHEN q1.q1_shares IS NULL THEN NULL
                        ELSE -100.0
                    END as change_pct,
                    CASE WHEN q1.q1_shares IS NULL THEN true ELSE false END as is_entry,
                    CASE WHEN q4.q4_shares IS NULL THEN true ELSE false END as is_exit
                FROM q4_agg q4
                FULL OUTER JOIN q1_agg q1 ON q4.cik = q1.cik
            )
            SELECT * FROM combined
            ORDER BY parent_name, ABS(change_shares) DESC
        """, [ticker, cusip, ticker, cusip]).fetchdf()

        results = []
        # Main holdings (top parents, not entries/exits)
        q2_top = q2[q2['parent_name'].isin(top_parents) & ~q2['is_entry'] & ~q2['is_exit']]

        # Count distinct funds per parent
        parent_child_counts = q2_top.groupby('parent_name')['fund_name'].nunique()

        current_parent = None
        for _, row in q2_top.iterrows():
            if row['parent_name'] != current_parent:
                current_parent = row['parent_name']
                parent_rows = q2_top[q2_top['parent_name'] == current_parent]
                child_count = int(parent_child_counts.get(current_parent, 1))
                p_q1 = parent_rows['q1_shares'].sum()
                p_q4 = parent_rows['q4_shares'].sum()
                p_chg = p_q4 - p_q1
                p_pct = (p_chg / p_q1 * 100) if p_q1 > 0 else None

                if child_count < 2:
                    # Collapse: single flat row with parent name and consolidated data
                    results.append({
                        'institution': current_parent,
                        'fund_name': current_parent,
                        'q1_shares': p_q1,
                        'q4_shares': p_q4,
                        'change_shares': p_chg,
                        'change_pct': p_pct,
                        'type': row['type'],
                        'is_parent': False,
                        'child_count': 1,
                        'section': 'holders',
                        'level': 0,
                    })
                else:
                    # Expand: parent summary row + child rows below
                    results.append({
                        'institution': current_parent,
                        'fund_name': '(parent total)',
                        'q1_shares': p_q1,
                        'q4_shares': p_q4,
                        'change_shares': p_chg,
                        'change_pct': p_pct,
                        'type': None,
                        'is_parent': True,
                        'child_count': child_count,
                        'section': 'holders',
                        'level': 0,
                    })
            # Only emit child rows if parent has 2+ children
            child_count = int(parent_child_counts.get(row['parent_name'], 1))
            if child_count >= 2:
                results.append({
                    'institution': row['parent_name'],
                    'fund_name': row['fund_name'],
                    'q1_shares': row['q1_shares'],
                    'q4_shares': row['q4_shares'],
                    'change_shares': row['change_shares'],
                    'change_pct': row['change_pct'],
                    'type': row['type'],
                    'is_parent': False,
                    'child_count': 0,
                    'section': 'holders',
                    'level': 1,
                })

        # Entries (new in Q4, >100K shares)
        entries = q2[q2['is_entry'] & (q2['q4_shares'] >= 100000)].sort_values(
            'q4_shares', ascending=False
        )
        for _, e in entries.head(15).iterrows():
            results.append({
                'institution': e['parent_name'],
                'fund_name': e['fund_name'],
                'q1_shares': None,
                'q4_shares': e['q4_shares'],
                'change_shares': e['q4_shares'],
                'change_pct': None,
                'type': e['type'],
                'is_parent': False,
                'child_count': 0,
                'section': 'entries',
                'level': 0,
            })

        # Exits (in Q1, gone in Q4, >100K shares)
        exits = q2[q2['is_exit'] & (q2['q1_shares'] >= 100000)].sort_values(
            'q1_shares', ascending=False
        )
        for _, e in exits.head(15).iterrows():
            results.append({
                'institution': e['parent_name'],
                'fund_name': e['fund_name'],
                'q1_shares': e['q1_shares'],
                'q4_shares': None,
                'change_shares': -e['q1_shares'] if pd.notna(e['q1_shares']) else None,
                'change_pct': -100.0,
                'type': e['type'],
                'is_parent': False,
                'child_count': 0,
                'section': 'exits',
                'level': 0,
            })
        return results
    finally:
        con.close()


def query3(ticker):
    """Active holder market cap analysis."""
    con = get_db()
    try:
        cusip = get_cusip(con, ticker)
        mktcap_row = con.execute(
            "SELECT market_cap FROM market_data WHERE ticker = ?", [ticker]
        ).fetchone()
        target_mktcap = mktcap_row[0] if mktcap_row else 0

        df = con.execute(f"""
            WITH cik_agg AS (
                SELECT
                    h.cik,
                    MAX(h.manager_name) as manager_name,
                    MAX(h.manager_type) as manager_type,
                    SUM(h.market_value_live) as position_value,
                    SUM(h.shares) as shares,
                    MAX(h.pct_of_portfolio) as pct_of_portfolio,
                    SUM(h.pct_of_float) as pct_of_float
                FROM holdings h
                WHERE h.quarter = '2025Q4'
                  AND (h.ticker = ? OR h.cusip = ?)
                  AND h.manager_type IN ('active', 'hedge_fund', 'activist', 'quantitative')
                GROUP BY h.cik
                ORDER BY position_value DESC NULLS LAST
                LIMIT 15
            ),
            with_percentile AS (
                SELECT
                    ca.*,
                    (
                        SELECT COUNT(*)
                        FROM holdings h2
                        INNER JOIN market_data m2 ON h2.ticker = m2.ticker
                        WHERE h2.cik = ca.cik AND h2.quarter = '2025Q4'
                          AND h2.security_type_inferred IN ('equity', 'etf')
                          AND m2.market_cap IS NOT NULL AND m2.market_cap > 0
                          AND m2.market_cap <= {target_mktcap}
                    ) as holdings_below,
                    (
                        SELECT COUNT(*)
                        FROM holdings h2
                        INNER JOIN market_data m2 ON h2.ticker = m2.ticker
                        WHERE h2.cik = ca.cik AND h2.quarter = '2025Q4'
                          AND h2.security_type_inferred IN ('equity', 'etf')
                          AND m2.market_cap IS NOT NULL AND m2.market_cap > 0
                    ) as total_with_mktcap
                FROM cik_agg ca
            )
            SELECT
                manager_name,
                position_value,
                pct_of_portfolio,
                pct_of_float,
                CASE WHEN total_with_mktcap > 0
                     THEN ROUND(holdings_below * 100.0 / total_with_mktcap, 1)
                     ELSE NULL END as mktcap_percentile,
                manager_type,
                '13F estimate' as source
            FROM with_percentile
            ORDER BY position_value DESC NULLS LAST
        """, [ticker, cusip]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query4(ticker):
    """Passive vs active ownership split."""
    con = get_db()
    try:
        df = con.execute("""
            SELECT
                CASE
                    WHEN manager_type = 'passive' THEN 'Passive (Index)'
                    WHEN manager_type = 'activist' THEN 'Activist'
                    WHEN manager_type IN ('active', 'hedge_fund', 'quantitative') THEN 'Active'
                    ELSE 'Other/Unknown'
                END as category,
                COUNT(DISTINCT cik) as num_holders,
                SUM(shares) as total_shares,
                SUM(market_value_live) as total_value,
                SUM(pct_of_float) as total_pct_float
            FROM holdings
            WHERE quarter = '2025Q4' AND ticker = ?
            GROUP BY category
            ORDER BY total_value DESC NULLS LAST
        """, [ticker]).fetchdf()
        grand_total = df['total_value'].sum()
        df['pct_of_inst'] = df['total_value'] / grand_total * 100 if grand_total > 0 else 0
        return df_to_records(df)
    finally:
        con.close()


def query5(ticker):
    """Quarterly share change heatmap."""
    con = get_db()
    try:
        df = con.execute("""
            WITH pivoted AS (
                SELECT
                    COALESCE(inst_parent_name, manager_name) as holder,
                    manager_type,
                    SUM(CASE WHEN quarter='2025Q1' THEN shares END) as q1_shares,
                    SUM(CASE WHEN quarter='2025Q2' THEN shares END) as q2_shares,
                    SUM(CASE WHEN quarter='2025Q3' THEN shares END) as q3_shares,
                    SUM(CASE WHEN quarter='2025Q4' THEN shares END) as q4_shares
                FROM holdings
                WHERE ticker = ?
                GROUP BY holder, manager_type
            )
            SELECT *,
                q2_shares - q1_shares as q1_to_q2,
                q3_shares - q2_shares as q2_to_q3,
                q4_shares - q3_shares as q3_to_q4,
                q4_shares - q1_shares as full_year_change
            FROM pivoted
            WHERE q4_shares IS NOT NULL
            ORDER BY q4_shares DESC
            LIMIT 25
        """, [ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query6(ticker):
    """Activist ownership tracker."""
    con = get_db()
    try:
        df = con.execute("""
            SELECT
                manager_name,
                quarter,
                shares,
                market_value_usd,
                market_value_live,
                pct_of_portfolio,
                pct_of_float
            FROM holdings
            WHERE ticker = ? AND is_activist = true
            ORDER BY manager_name, quarter
        """, [ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query7(ticker, cik=None, fund_name=None):
    """Single fund portfolio — aggregated by ticker, with stats header."""
    con = get_db()
    try:
        if not cik:
            # Default to top non-passive holder of the ticker
            row = con.execute("""
                SELECT cik, fund_name FROM holdings
                WHERE ticker = ? AND quarter = '2025Q4'
                  AND manager_type NOT IN ('passive')
                ORDER BY market_value_live DESC NULLS LAST
                LIMIT 1
            """, [ticker]).fetchone()
            if not row:
                return {'stats': {}, 'positions': []}
            cik = row[0]
            fund_name = fund_name or row[1]

        # Build the WHERE filter — cik always, fund_name when provided
        where = "h.cik = ? AND h.quarter = '2025Q4'"
        params = [cik]
        if fund_name:
            where += " AND h.fund_name = ?"
            params.append(fund_name)

        # Fund metadata
        meta_where = "cik = ? AND quarter = '2025Q4'"
        meta_params = [cik]
        if fund_name:
            meta_where += " AND fund_name = ?"
            meta_params.append(fund_name)
        mgr_row = con.execute(f"""
            SELECT fund_name, MAX(manager_type) as manager_type
            FROM holdings WHERE {meta_where}
            GROUP BY fund_name LIMIT 1
        """, meta_params).fetchone()
        display_name = mgr_row[0] if mgr_row else cik
        mgr_type = mgr_row[1] if mgr_row else 'unknown'

        # Aggregated portfolio by ticker
        df = con.execute(f"""
            SELECT
                h.ticker,
                MAX(h.issuer_name) as issuer_name,
                MAX(s.sector) as sector,
                SUM(h.shares) as shares,
                SUM(h.market_value_live) as market_value_live,
                MAX(h.pct_of_portfolio) as pct_of_portfolio,
                SUM(h.pct_of_float) as pct_of_float,
                MAX(m.market_cap) as market_cap
            FROM holdings h
            LEFT JOIN market_data m ON h.ticker = m.ticker
            LEFT JOIN (
                SELECT cusip, MAX(sector) as sector
                FROM securities WHERE sector IS NOT NULL AND sector != ''
                GROUP BY cusip
            ) s ON h.cusip = s.cusip
            WHERE {where}
            GROUP BY h.ticker
            ORDER BY market_value_live DESC NULLS LAST
        """, params).fetchdf()

        records = df_to_records(df)

        # Add rank
        for i, r in enumerate(records, 1):
            r['rank'] = i

        # Portfolio stats
        total_value = df['market_value_live'].sum()
        num_positions = len(df)
        top10_value = df.head(10)['market_value_live'].sum()
        top10_pct = (top10_value / total_value * 100) if total_value > 0 else 0

        stats = {
            'manager_name': display_name,
            'cik': cik,
            'manager_type': mgr_type,
            'total_value': total_value,
            'num_positions': num_positions,
            'top10_concentration_pct': round(top10_pct, 2),
        }

        return clean_for_json({'stats': stats, 'positions': records})
    finally:
        con.close()


def query8(ticker):
    """Cross-holder overlap — stocks most commonly held by same institutions."""
    con = get_db()
    try:
        df = con.execute("""
            WITH target_holders AS (
                SELECT DISTINCT cik
                FROM holdings
                WHERE ticker = ? AND quarter = '2025Q4'
            )
            SELECT
                h.ticker,
                h.issuer_name,
                COUNT(DISTINCT h.cik) as shared_holders,
                SUM(h.market_value_live) as total_value,
                (SELECT COUNT(*) FROM target_holders) as target_holders_count
            FROM holdings h
            INNER JOIN target_holders th ON h.cik = th.cik
            WHERE h.quarter = '2025Q4'
              AND h.ticker != ?
              AND h.ticker IS NOT NULL
            GROUP BY h.ticker, h.issuer_name
            ORDER BY shared_holders DESC
            LIMIT 20
        """, [ticker, ticker]).fetchdf()
        if len(df) > 0:
            df['overlap_pct'] = df['shared_holders'] / df['target_holders_count'] * 100
        return df_to_records(df)
    finally:
        con.close()


def query9(ticker):
    """Sector rotation analysis — sector allocation of active holders."""
    con = get_db()
    try:
        df = con.execute("""
            WITH target_ciks AS (
                SELECT DISTINCT cik
                FROM holdings
                WHERE ticker = ? AND quarter = '2025Q4'
                AND manager_type IN ('active', 'hedge_fund')
            )
            SELECT
                s.sector,
                COUNT(DISTINCT h.ticker) as num_stocks,
                SUM(h.market_value_live) as sector_value,
                SUM(h.market_value_live) * 100.0 / SUM(SUM(h.market_value_live)) OVER () as pct_of_total
            FROM holdings h
            INNER JOIN target_ciks tc ON h.cik = tc.cik
            INNER JOIN securities s ON h.cusip = s.cusip
            WHERE h.quarter = '2025Q4' AND s.sector IS NOT NULL AND s.sector != ''
            GROUP BY s.sector
            ORDER BY sector_value DESC NULLS LAST
        """, [ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query10(ticker):
    """Largest new positions (Q4 entries)."""
    con = get_db()
    try:
        df = con.execute("""
            SELECT
                q4.manager_name,
                q4.manager_type,
                q4.shares,
                q4.market_value_live,
                q4.pct_of_portfolio,
                q4.pct_of_float
            FROM holdings q4
            LEFT JOIN holdings q3 ON q4.cik = q3.cik AND q3.ticker = ? AND q3.quarter = '2025Q3'
            WHERE q4.ticker = ? AND q4.quarter = '2025Q4' AND q3.cik IS NULL
            ORDER BY q4.market_value_live DESC NULLS LAST
            LIMIT 20
        """, [ticker, ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query11(ticker):
    """Largest exits (Q3 holders gone in Q4)."""
    con = get_db()
    try:
        df = con.execute("""
            SELECT
                q3.manager_name,
                q3.manager_type,
                q3.shares as q3_shares,
                q3.market_value_usd as q3_value,
                q3.pct_of_portfolio as q3_pct
            FROM holdings q3
            LEFT JOIN holdings q4 ON q3.cik = q4.cik AND q4.ticker = ? AND q4.quarter = '2025Q4'
            WHERE q3.ticker = ? AND q3.quarter = '2025Q3' AND q4.cik IS NULL
            ORDER BY q3.market_value_usd DESC
            LIMIT 20
        """, [ticker, ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query12(ticker):
    """Concentration analysis — top holders cumulative % of float."""
    con = get_db()
    try:
        df = con.execute("""
            WITH ranked AS (
                SELECT
                    COALESCE(inst_parent_name, manager_name) as holder,
                    SUM(pct_of_float) as total_pct_float,
                    SUM(shares) as total_shares,
                    ROW_NUMBER() OVER (ORDER BY SUM(pct_of_float) DESC) as rn
                FROM holdings
                WHERE ticker = ? AND quarter = '2025Q4' AND pct_of_float IS NOT NULL
                GROUP BY holder
            )
            SELECT
                rn as rank,
                holder,
                total_pct_float,
                total_shares,
                SUM(total_pct_float) OVER (ORDER BY rn) as cumulative_pct
            FROM ranked
            ORDER BY rn
            LIMIT 20
        """, [ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query13(ticker=None):
    """Energy sector institutional rotation (Q1 to Q4 2025)."""
    con = get_db()
    try:
        df = con.execute("""
            WITH energy_moves AS (
                SELECT
                    h4.ticker,
                    h4.issuer_name,
                    COUNT(DISTINCT CASE WHEN h4.shares > COALESCE(h1.shares, 0) THEN h4.cik END) as buyers,
                    COUNT(DISTINCT CASE WHEN h4.shares < COALESCE(h1.shares, 0) THEN h4.cik END) as sellers,
                    COUNT(DISTINCT CASE WHEN h1.cik IS NULL THEN h4.cik END) as new_positions,
                    SUM(h4.market_value_live) as q4_total_value
                FROM holdings h4
                INNER JOIN securities s ON h4.cusip = s.cusip AND s.is_energy = true
                LEFT JOIN holdings h1 ON h4.cik = h1.cik AND h4.ticker = h1.ticker AND h1.quarter = '2025Q1'
                WHERE h4.quarter = '2025Q4'
                  AND h4.manager_type IN ('active', 'hedge_fund', 'activist')
                GROUP BY h4.ticker, h4.issuer_name
            )
            SELECT *,
                buyers - sellers as net_flow,
                ROUND(buyers * 100.0 / (buyers + sellers), 1) as buy_pct
            FROM energy_moves
            WHERE buyers + sellers >= 5
            ORDER BY net_flow DESC
            LIMIT 25
        """).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query14(ticker):
    """Manager AUM vs position size."""
    con = get_db()
    try:
        df = con.execute("""
            SELECT
                h.manager_name,
                h.manager_type,
                h.is_activist,
                m.aum_total / 1e9 as manager_aum_bn,
                h.market_value_live / 1e6 as position_mm,
                h.pct_of_portfolio,
                h.shares
            FROM holdings h
            LEFT JOIN managers m ON h.cik = m.cik
            WHERE h.ticker = ? AND h.quarter = '2025Q4'
              AND m.aum_total IS NOT NULL AND m.aum_total > 0
            ORDER BY h.market_value_live DESC NULLS LAST
            LIMIT 50
        """, [ticker]).fetchdf()
        return df_to_records(df)
    finally:
        con.close()


def query15(ticker=None):
    """Database statistics."""
    con = get_db()
    try:
        stats = {}
        stats['total_holdings'] = con.execute('SELECT COUNT(*) FROM holdings').fetchone()[0]
        stats['unique_filers'] = con.execute('SELECT COUNT(DISTINCT cik) FROM holdings').fetchone()[0]
        stats['unique_securities'] = con.execute('SELECT COUNT(DISTINCT cusip) FROM holdings').fetchone()[0]
        stats['quarters_loaded'] = con.execute('SELECT COUNT(DISTINCT quarter) FROM holdings').fetchone()[0]
        stats['manager_records'] = con.execute('SELECT COUNT(*) FROM managers').fetchone()[0]
        stats['securities_mapped'] = con.execute('SELECT COUNT(*) FROM securities').fetchone()[0]
        stats['market_data_tickers'] = con.execute('SELECT COUNT(*) FROM market_data').fetchone()[0]
        stats['adv_records'] = con.execute('SELECT COUNT(*) FROM adv_managers').fetchone()[0]

        # Quarter breakdown
        qstats = con.execute("""
            SELECT quarter, COUNT(*) as rows, COUNT(DISTINCT cik) as filers,
                   COUNT(DISTINCT cusip) as securities,
                   SUM(market_value_usd) / 1e12 as total_value_tn
            FROM holdings GROUP BY quarter ORDER BY quarter
        """).fetchdf()
        stats['quarters'] = df_to_records(qstats)

        # Coverage rates
        coverage = con.execute("""
            SELECT
                COUNT(*) as total,
                COUNT(CASE WHEN ticker IS NOT NULL THEN 1 END) as with_ticker,
                COUNT(CASE WHEN manager_type IS NOT NULL THEN 1 END) as with_manager_type,
                COUNT(CASE WHEN market_value_live IS NOT NULL THEN 1 END) as with_live_value,
                COUNT(CASE WHEN pct_of_float IS NOT NULL THEN 1 END) as with_float_pct
            FROM holdings WHERE quarter = '2025Q4'
        """).fetchone()
        total = coverage[0] or 1
        stats['coverage'] = {
            'total': total,
            'ticker_pct': round(coverage[1] / total * 100, 1),
            'manager_type_pct': round(coverage[2] / total * 100, 1),
            'live_value_pct': round(coverage[3] / total * 100, 1),
            'float_pct_pct': round(coverage[4] / total * 100, 1),
        }
        return clean_for_json([stats])
    finally:
        con.close()


# Map query number to function
QUERY_FUNCTIONS = {
    1: query1, 2: query2, 3: query3, 4: query4, 5: query5,
    6: query6, 7: query7, 8: query8, 9: query9, 10: query10,
    11: query11, 12: query12, 13: query13, 14: query14, 15: query15,
}

QUERY_NAMES = {
    1: 'Register', 2: '4Q Change', 3: 'Active Analysis', 4: 'Passive/Active Split',
    5: 'Quarterly Trend', 6: 'Activist Tracker', 7: 'Fund Portfolio',
    8: 'Cross-Ownership', 9: 'Sector Rotation', 10: 'New Positions',
    11: 'Exits', 12: 'Concentration', 13: 'Energy Rotation',
    14: 'AUM vs Position', 15: 'DB Statistics',
}


# ---------------------------------------------------------------------------
# Summary endpoint
# ---------------------------------------------------------------------------

def get_summary(ticker):
    """Quick summary stats for the header card."""
    con = get_db()
    try:
        cusip = get_cusip(con, ticker)
        if not cusip:
            return None

        # Company name — use most common issuer_name from filings (avoids CUSIP cross-contamination)
        name_row = con.execute(
            "SELECT MODE(issuer_name) FROM holdings WHERE ticker = ? AND quarter = '2025Q4'",
            [ticker]
        ).fetchone()
        company_name = name_row[0] if name_row else ticker

        # Latest quarter
        q_row = con.execute("""
            SELECT MAX(quarter) FROM holdings WHERE ticker = ?
        """, [ticker]).fetchone()
        latest_quarter = q_row[0] if q_row else 'N/A'

        # Total institutional holdings
        totals = con.execute("""
            SELECT
                SUM(market_value_live) as total_value,
                SUM(pct_of_float) as total_pct_float,
                COUNT(DISTINCT cik) as num_holders,
                SUM(shares) as total_shares
            FROM holdings
            WHERE ticker = ? AND quarter = '2025Q4'
        """, [ticker]).fetchone()

        # Active vs passive split
        split = con.execute("""
            SELECT
                SUM(CASE WHEN manager_type = 'passive' THEN market_value_live ELSE 0 END) as passive_value,
                SUM(CASE WHEN manager_type IN ('active','hedge_fund','quantitative','activist')
                    THEN market_value_live ELSE 0 END) as active_value
            FROM holdings
            WHERE ticker = ? AND quarter = '2025Q4'
        """, [ticker]).fetchone()

        # Market data
        mkt = con.execute(
            "SELECT price_live, market_cap, float_shares FROM market_data WHERE ticker = ?",
            [ticker]
        ).fetchone()

        return {
            'company_name': company_name,
            'ticker': ticker,
            'latest_quarter': latest_quarter,
            'total_value': totals[0],
            'total_pct_float': totals[1],
            'num_holders': totals[2],
            'total_shares': totals[3],
            'passive_value': split[0] if split else None,
            'active_value': split[1] if split else None,
            'price': mkt[0] if mkt else None,
            'market_cap': mkt[1] if mkt else None,
            'shares_float': mkt[2] if mkt else None,
        }
        return clean_for_json(result)
    finally:
        con.close()


# ---------------------------------------------------------------------------
# Excel export helper
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color='002147', end_color='002147', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
ALT_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
DATA_FONT = Font(name='Arial', size=10)


def build_excel(data, sheet_name='Data'):
    """Build a formatted .xlsx workbook from a list of dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit

    if not data:
        ws.append(['No data'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # Flatten nested dicts for export
    flat_data = []
    for row in data:
        flat = {}
        for k, v in row.items():
            if k.startswith('_'):
                continue
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat[f"{k}_{k2}"] = v2
            else:
                flat[k] = v
        flat_data.append(flat)

    headers = list(flat_data[0].keys())
    ws.append(headers)

    # Style header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, record in enumerate(flat_data, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = record.get(key)
            cell.value = val
            cell.font = DATA_FONT

            # Number formatting
            if isinstance(val, (int, float)) and val is not None:
                if 'pct' in key.lower() or 'percent' in key.lower():
                    cell.number_format = '0.00"%"'
                elif 'value' in key.lower() or 'aum' in key.lower() or 'cap' in key.lower():
                    cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

        # Alternating row fill
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL

    # Auto-fit column widths
    for col_idx, key in enumerate(headers, 1):
        max_len = len(str(key))
        for row_idx in range(2, min(len(flat_data) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    from flask import render_template
    return render_template('index.html')


@app.route('/api/tickers')
def api_tickers():
    try:
        con = get_db()
    except Exception as e:
        return jsonify({'error': f'Database unavailable: {e}'}), 503
    try:
        df = con.execute("""
            SELECT ticker, MODE(issuer_name) as name
            FROM holdings
            WHERE ticker IS NOT NULL AND ticker != '' AND quarter = '2025Q4'
            GROUP BY ticker
            ORDER BY ticker
        """).fetchdf()
        return jsonify(df_to_records(df))
    finally:
        con.close()


@app.route('/api/fund_portfolio_managers')
def api_fund_portfolio_managers():
    ticker = request.args.get('ticker', '').upper().strip()
    if not ticker:
        return jsonify({'error': 'Missing ticker parameter'}), 400
    try:
        con = get_db()
    except Exception as e:
        return jsonify({'error': f'Database unavailable: {e}'}), 503
    try:
        df = con.execute("""
            SELECT
                cik,
                fund_name,
                MAX(COALESCE(inst_parent_name, manager_name)) as inst_parent_name,
                SUM(market_value_live) as position_value,
                MAX(manager_type) as manager_type
            FROM holdings
            WHERE ticker = ? AND quarter = '2025Q4'
              AND manager_type NOT IN ('passive')
            GROUP BY cik, fund_name
            ORDER BY position_value DESC NULLS LAST
            LIMIT 50
        """, [ticker]).fetchdf()
        return jsonify(df_to_records(df))
    finally:
        con.close()


@app.route('/api/summary')
def api_summary():
    ticker = request.args.get('ticker', '').upper().strip()
    if not ticker:
        return jsonify({'error': 'Missing ticker parameter'}), 400
    result = get_summary(ticker)
    if not result:
        return jsonify({'error': f'No data found for ticker {ticker}'}), 404
    return jsonify(result)


@app.route('/api/query<int:qnum>')
def api_query(qnum):
    if qnum not in QUERY_FUNCTIONS:
        return jsonify({'error': f'Invalid query number: {qnum}'}), 400
    ticker = request.args.get('ticker', '').upper().strip()
    cik = request.args.get('cik', '').strip()

    # Queries 13 and 15 do not require a ticker
    if qnum not in (13, 15) and not ticker:
        return jsonify({'error': 'Missing ticker parameter'}), 400

    try:
        fn = QUERY_FUNCTIONS[qnum]
        if qnum == 7:
            fund_name = request.args.get('fund_name', '').strip() or None
            data = fn(ticker, cik=cik or None, fund_name=fund_name)
            # query7 returns {stats, positions} dict
            if not data.get('positions'):
                return jsonify({'error': f'No holdings found for CIK {cik}'}), 404
            return jsonify(data)
        elif qnum in (13, 15):
            data = fn(ticker or None)
        else:
            data = fn(ticker)

        if not data:
            return jsonify({'error': f'No data found for ticker {ticker}'}), 404
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/query<int:qnum>')
def api_export(qnum):
    if qnum not in QUERY_FUNCTIONS:
        return jsonify({'error': f'Invalid query number: {qnum}'}), 400
    ticker = request.args.get('ticker', '').upper().strip()
    cik = request.args.get('cik', '').strip()

    if qnum not in (13, 15) and not ticker:
        return jsonify({'error': 'Missing ticker parameter'}), 400

    try:
        fn = QUERY_FUNCTIONS[qnum]
        if qnum == 7 and cik:
            data = fn(ticker or 'AR', cik=cik)
        elif qnum in (13, 15):
            data = fn(ticker or None)
        else:
            data = fn(ticker)

        if not data:
            return jsonify({'error': f'No data found for ticker {ticker}'}), 404

        qname = QUERY_NAMES.get(qnum, f'Query{qnum}')
        sheet_name = f"{qname} - {ticker or 'ALL'}"
        buf = build_excel(data, sheet_name=sheet_name)

        filename = f"query{qnum}_{ticker or 'ALL'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='13F Ownership Research Web App')
    parser.add_argument('--port', type=int, default=8001, help='Port to run on (default: 8001)')
    args = parser.parse_args()

    # Check PORT env var (for Render deployment)
    port = int(os.environ.get('PORT', args.port))

    # Resolve database path at startup (creates snapshot if main DB is locked)
    _init_db_path()

    print()
    print('  13F Ownership Research')
    print(f'  Database: {_active_db_path}')
    if _active_db_path != DB_PATH:
        print(f'  (main DB locked — serving from snapshot)')
    print(f'  Running at: http://localhost:{port}')
    print()

    app.run(host='0.0.0.0', port=port, debug=False)
